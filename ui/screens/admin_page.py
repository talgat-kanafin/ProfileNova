"""
ui/screens/admin_page.py
=========================
Раздел "Администрирование" — только для role=admin.

Вкладка 1: Пользователи
  — список всех пользователей
  — создание / редактирование / отключение
  — смена пароля

Вкладка 2: Доступ к проектам
  — выбрать проект → назначить/снять доступ пользователям

Вкладка 3: Журнал изменений
  — фильтр по проекту / пользователю / дате
  — экспорт в Excel
"""

import os, sys
from datetime import datetime
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QTabWidget, QFrame, QPushButton, QTableWidget,
    QTableWidgetItem, QHeaderView, QComboBox,
    QLineEdit, QCheckBox, QDialog, QDialogButtonBox,
    QMessageBox, QRadioButton, QButtonGroup,
    QAbstractItemView, QSplitter, QListWidget,
    QListWidgetItem, QFormLayout
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QColor

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from ui.theme import Colors, Typography, Spacing, Sizes
from ui.i18n  import tr


class AdminPage(QWidget):

    def __init__(self, ctx, parent=None):
        super().__init__(parent)
        self._ctx = ctx
        self._build_ui()
        self._load_all()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        self._tabs = QTabWidget()
        self._tabs.setStyleSheet(f"""
            QTabWidget::pane {{ border: none; background-color: {Colors.BG_BASE}; }}
            QTabBar::tab {{
                background-color: {Colors.BG_SURFACE};
                color: {Colors.TEXT_SECONDARY};
                padding: 10px 24px; border: none;
                border-bottom: 2px solid transparent; font-size: 13px;
            }}
            QTabBar::tab:selected {{
                color: {Colors.TEXT_PRIMARY};
                border-bottom: 2px solid {Colors.ACCENT};
            }}
            QTabBar::tab:hover {{
                color: {Colors.TEXT_PRIMARY};
                background-color: {Colors.BG_ELEVATED};
            }}
        """)

        self._tabs.addTab(self._build_users_tab(),   "👤  Пользователи")
        self._tabs.addTab(self._build_access_tab(),  "🔑  Доступ")
        self._tabs.addTab(self._build_audit_tab(),   "📋  Журнал изменений")

        root.addWidget(self._tabs)

    # ════════════════════════════════════════════════════════
    # ВКЛАДКА: ПОЛЬЗОВАТЕЛИ
    # ════════════════════════════════════════════════════════

    def _build_users_tab(self) -> QWidget:
        w = QWidget()
        w.setStyleSheet(f"background-color: {Colors.BG_BASE};")
        layout = QVBoxLayout(w)
        layout.setContentsMargins(Spacing.XL, Spacing.XL, Spacing.XL, Spacing.XL)
        layout.setSpacing(Spacing.MD)

        # Тулбар
        tb = QHBoxLayout()
        title = QLabel("Пользователи системы")
        title.setFont(Typography.get_font(14, bold=True))
        tb.addWidget(title)
        tb.addStretch()

        self._user_search = QLineEdit()
        self._user_search.setPlaceholderText("Поиск...")
        self._user_search.setFixedWidth(200)
        self._user_search.setFixedHeight(32)
        self._user_search.textChanged.connect(self._filter_users)
        tb.addWidget(self._user_search)

        btn_new = QPushButton("＋  Новый пользователь")
        btn_new.setFixedHeight(32)
        btn_new.setStyleSheet(f"""
            QPushButton {{
                background-color: {Colors.ACCENT}; color: white;
                border: none; border-radius: 4px; font-weight: 600; padding: 0 12px;
            }}
            QPushButton:hover {{ background-color: {Colors.ACCENT_HOVER}; }}
        """)
        btn_new.clicked.connect(self._new_user)
        tb.addWidget(btn_new)
        layout.addLayout(tb)

        # Таблица пользователей
        self._users_table = QTableWidget(0, 6)
        self._users_table.setHorizontalHeaderLabels([
            "Логин", "ФИО", "Подразделение", "Роль", "Последний вход", "Статус"
        ])
        self._users_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch
        )
        self._users_table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.Stretch
        )
        self._users_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._users_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._users_table.setAlternatingRowColors(True)
        self._users_table.doubleClicked.connect(self._edit_user)
        self._users_table.setStyleSheet(self._table_style())
        layout.addWidget(self._users_table, 1)

        # Кнопки действий
        btn_row = QHBoxLayout()
        btn_edit = QPushButton("✏  Редактировать")
        btn_edit.setFixedHeight(32)
        btn_edit.clicked.connect(self._edit_user)
        btn_row.addWidget(btn_edit)

        btn_pwd = QPushButton("🔒  Сменить пароль")
        btn_pwd.setFixedHeight(32)
        btn_pwd.clicked.connect(self._change_password)
        btn_row.addWidget(btn_pwd)

        btn_toggle = QPushButton("⏸  Отключить / Включить")
        btn_toggle.setFixedHeight(32)
        btn_toggle.clicked.connect(self._toggle_user)
        btn_row.addWidget(btn_toggle)

        btn_row.addStretch()
        layout.addLayout(btn_row)
        return w

    # ════════════════════════════════════════════════════════
    # ВКЛАДКА: ДОСТУП К ПРОЕКТАМ
    # ════════════════════════════════════════════════════════

    def _build_access_tab(self) -> QWidget:
        w = QWidget()
        w.setStyleSheet(f"background-color: {Colors.BG_BASE};")
        layout = QVBoxLayout(w)
        layout.setContentsMargins(Spacing.XL, Spacing.XL, Spacing.XL, Spacing.XL)
        layout.setSpacing(Spacing.MD)

        # Выбор проекта
        top = QHBoxLayout()
        top.addWidget(QLabel("Проект:"))
        self._access_project = QComboBox()
        self._access_project.setFixedWidth(300)
        self._access_project.setFixedHeight(32)
        self._access_project.currentIndexChanged.connect(self._load_project_access)
        top.addWidget(self._access_project)
        top.addStretch()
        layout.addLayout(top)

        # Сплиттер: все пользователи | с доступом
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # Левая: все пользователи
        left = QFrame()
        left.setStyleSheet(f"""
            background-color: {Colors.BG_SURFACE};
            border: 1px solid {Colors.BORDER}; border-radius: 6px;
        """)
        ll = QVBoxLayout(left)
        ll.setContentsMargins(Spacing.MD, Spacing.MD, Spacing.MD, Spacing.MD)
        ll.addWidget(self._section_label("Все пользователи"))
        self._all_users_list = QListWidget()
        self._all_users_list.setStyleSheet(self._list_style())
        self._all_users_list.setSelectionMode(
            QAbstractItemView.SelectionMode.MultiSelection
        )
        ll.addWidget(self._all_users_list)
        splitter.addWidget(left)

        # Центр: кнопки
        center = QWidget()
        center.setFixedWidth(120)
        cl = QVBoxLayout(center)
        cl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        cl.setSpacing(Spacing.SM)

        btn_grant_edit = QPushButton("▶ Редактор")
        btn_grant_edit.setFixedHeight(36)
        btn_grant_edit.setStyleSheet(f"""
            QPushButton {{
                background-color: {Colors.ACCENT}; color: white;
                border: none; border-radius: 4px; font-size: 12px;
            }}
            QPushButton:hover {{ background-color: {Colors.ACCENT_HOVER}; }}
        """)
        btn_grant_edit.clicked.connect(lambda: self._grant_access(can_edit=True))
        cl.addWidget(btn_grant_edit)

        btn_grant_view = QPushButton("▶ Просмотр")
        btn_grant_view.setFixedHeight(36)
        btn_grant_view.clicked.connect(lambda: self._grant_access(can_edit=False))
        cl.addWidget(btn_grant_view)

        btn_revoke = QPushButton("◀ Убрать")
        btn_revoke.setFixedHeight(36)
        btn_revoke.clicked.connect(self._revoke_access)
        cl.addWidget(btn_revoke)

        splitter.addWidget(center)

        # Правая: пользователи с доступом
        right = QFrame()
        right.setStyleSheet(left.styleSheet())
        rl = QVBoxLayout(right)
        rl.setContentsMargins(Spacing.MD, Spacing.MD, Spacing.MD, Spacing.MD)
        rl.addWidget(self._section_label("Имеют доступ к проекту"))
        self._access_list = QListWidget()
        self._access_list.setStyleSheet(self._list_style())
        self._access_list.setSelectionMode(
            QAbstractItemView.SelectionMode.MultiSelection
        )
        rl.addWidget(self._access_list)
        splitter.addWidget(right)

        splitter.setSizes([400, 120, 400])
        layout.addWidget(splitter, 1)
        return w

    # ════════════════════════════════════════════════════════
    # ВКЛАДКА: ЖУРНАЛ ИЗМЕНЕНИЙ
    # ════════════════════════════════════════════════════════

    def _build_audit_tab(self) -> QWidget:
        w = QWidget()
        w.setStyleSheet(f"background-color: {Colors.BG_BASE};")
        layout = QVBoxLayout(w)
        layout.setContentsMargins(Spacing.XL, Spacing.XL, Spacing.XL, Spacing.XL)
        layout.setSpacing(Spacing.MD)

        # Фильтры
        filter_frame = QFrame()
        filter_frame.setStyleSheet(f"""
            background-color: {Colors.BG_SURFACE};
            border: 1px solid {Colors.BORDER}; border-radius: 6px;
        """)
        ff = QHBoxLayout(filter_frame)
        ff.setContentsMargins(Spacing.MD, Spacing.SM, Spacing.MD, Spacing.SM)
        ff.setSpacing(Spacing.MD)

        ff.addWidget(self._section_label("Проект:"))
        self._audit_project = QComboBox()
        self._audit_project.setFixedWidth(200)
        self._audit_project.setFixedHeight(30)
        self._audit_project.addItem("Все проекты", None)
        ff.addWidget(self._audit_project)

        ff.addWidget(self._section_label("Пользователь:"))
        self._audit_user = QComboBox()
        self._audit_user.setFixedWidth(180)
        self._audit_user.setFixedHeight(30)
        self._audit_user.addItem("Все", None)
        ff.addWidget(self._audit_user)

        ff.addWidget(self._section_label("Таблица:"))
        self._audit_table_filter = QComboBox()
        self._audit_table_filter.setFixedWidth(160)
        self._audit_table_filter.setFixedHeight(30)
        self._audit_table_filter.addItem("Все таблицы", None)
        ff.addWidget(self._audit_table_filter)

        ff.addStretch()

        btn_filter = QPushButton("🔍 Применить")
        btn_filter.setFixedHeight(30)
        btn_filter.clicked.connect(self._load_audit)
        ff.addWidget(btn_filter)

        btn_export = QPushButton("📊 Excel")
        btn_export.setFixedHeight(30)
        btn_export.clicked.connect(self._export_audit)
        ff.addWidget(btn_export)

        layout.addWidget(filter_frame)

        # Таблица журнала
        self._audit_table = QTableWidget(0, 6)
        self._audit_table.setHorizontalHeaderLabels([
            "Дата/Время", "Пользователь", "Проект",
            "Таблица", "Действие", "Запись №"
        ])
        self._audit_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.ResizeToContents
        )
        self._audit_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch
        )
        self._audit_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._audit_table.setAlternatingRowColors(True)
        self._audit_table.setStyleSheet(self._table_style())
        layout.addWidget(self._audit_table, 1)

        self._audit_count = QLabel("")
        self._audit_count.setStyleSheet(
            f"color: {Colors.TEXT_SECONDARY}; font-size: 11px;"
        )
        layout.addWidget(self._audit_count)

        return w

    # ── Загрузка данных ──────────────────────────────────────

    def _load_all(self):
        self._load_users()
        self._load_projects_for_combos()
        self._load_audit()

    def _load_users(self):
        rows = self._ctx.db.fetchall(
            "SELECT * FROM users ORDER BY full_name"
        )
        self._users_table.setRowCount(len(rows))
        role_labels = {
            "admin":  "Администратор",
            "editor": "Редактор",
            "viewer": "Просмотр"
        }
        action_colors = {
            1: Colors.TEXT_PRIMARY,
            0: Colors.TEXT_DISABLED
        }
        for i, u in enumerate(rows):
            is_active = u.get("is_active", 1)
            cells = [
                u.get("username", ""),
                u.get("full_name", ""),
                u.get("department", "") or "",
                role_labels.get(u.get("role", ""), u.get("role", "")),
                (u.get("last_login") or "Никогда")[:16],
                "Активен" if is_active else "Отключён"
            ]
            for j, text in enumerate(cells):
                item = QTableWidgetItem(text)
                item.setData(Qt.ItemDataRole.UserRole, u["id"])
                if not is_active:
                    item.setForeground(QColor(Colors.TEXT_DISABLED))
                if j == 5 and not is_active:
                    item.setForeground(QColor(Colors.ERROR))
                self._users_table.setItem(i, j, item)

        # Список для вкладки доступа
        self._all_users_list.clear()
        for u in rows:
            dept = f" ({u['department']})" if u.get("department") else ""
            item = QListWidgetItem(f"{u['full_name']}{dept}")
            item.setData(Qt.ItemDataRole.UserRole, u["id"])
            self._all_users_list.addItem(item)

        # Фильтр журнала
        self._audit_user.clear()
        self._audit_user.addItem("Все", None)
        for u in rows:
            self._audit_user.addItem(u["full_name"], u["id"])

    def _load_projects_for_combos(self):
        projects = self._ctx.db.fetchall(
            "SELECT id, name FROM projects ORDER BY name"
        )
        for combo in (self._access_project, self._audit_project):
            combo.clear()
            if combo is self._audit_project:
                combo.addItem("Все проекты", None)
            for p in projects:
                combo.addItem(p["name"], p["id"])

        if self._access_project.count() > 0:
            self._load_project_access()

    def _load_project_access(self):
        project_id = self._access_project.currentData()
        if project_id is None:
            return

        has_access = self._ctx.db.fetchall(
            """
            SELECT u.id, u.full_name, u.department, pa.can_edit
            FROM project_access pa
            JOIN users u ON u.id = pa.user_id
            WHERE pa.project_id = ?
            ORDER BY u.full_name
            """,
            (project_id,)
        )
        access_ids = {r["id"] for r in has_access}

        self._access_list.clear()
        for r in has_access:
            dept    = f" ({r['department']})" if r.get("department") else ""
            role    = "✏ Редактор" if r["can_edit"] else "👁 Просмотр"
            item    = QListWidgetItem(f"{r['full_name']}{dept}  —  {role}")
            item.setData(Qt.ItemDataRole.UserRole, r["id"])
            if r["can_edit"]:
                item.setForeground(QColor(Colors.ACCENT))
            self._access_list.addItem(item)

        # Скрыть уже имеющих доступ в левом списке
        for i in range(self._all_users_list.count()):
            it  = self._all_users_list.item(i)
            uid = it.data(Qt.ItemDataRole.UserRole)
            it.setHidden(uid in access_ids)

    def _load_audit(self):
        project_id = self._audit_project.currentData()
        user_id    = self._audit_user.currentData()

        sql    = """
            SELECT a.changed_at, u.full_name, p.name as project_name,
                   a.table_name, a.action, a.record_id
            FROM audit_log a
            LEFT JOIN users    u ON u.id = a.user_id
            LEFT JOIN projects p ON p.id = a.project_id
            WHERE 1=1
        """
        params = []
        if project_id:
            sql += " AND a.project_id=?"
            params.append(project_id)
        if user_id:
            sql += " AND a.user_id=?"
            params.append(user_id)

        sql += " ORDER BY a.changed_at DESC LIMIT 500"

        rows = self._ctx.db.fetchall(sql, tuple(params))
        self._audit_table.setRowCount(len(rows))

        action_colors = {
            "INSERT": Colors.SUCCESS,
            "UPDATE": Colors.INFO,
            "DELETE": Colors.ERROR,
            "IMPORT": Colors.WARNING,
        }

        for i, r in enumerate(rows):
            cells = [
                (r.get("changed_at") or "")[:16],
                r.get("full_name")    or "—",
                r.get("project_name") or "—",
                r.get("table_name")   or "",
                r.get("action")       or "",
                str(r.get("record_id") or ""),
            ]
            for j, text in enumerate(cells):
                item = QTableWidgetItem(text)
                if j == 4:  # Действие
                    color = action_colors.get(text, Colors.TEXT_PRIMARY)
                    item.setForeground(QColor(color))
                self._audit_table.setItem(i, j, item)

        self._audit_count.setText(f"Записей: {len(rows)}")

    # ── Действия: Пользователи ───────────────────────────────

    def _filter_users(self, text: str):
        text = text.lower()
        for row in range(self._users_table.rowCount()):
            match = not text
            if not match:
                for col in range(self._users_table.columnCount()):
                    item = self._users_table.item(row, col)
                    if item and text in item.text().lower():
                        match = True
                        break
            self._users_table.setRowHidden(row, not match)

    def _get_selected_user_id(self) -> int | None:
        row = self._users_table.currentRow()
        if row < 0:
            return None
        item = self._users_table.item(row, 0)
        return item.data(Qt.ItemDataRole.UserRole) if item else None

    def _new_user(self):
        dlg = UserDialog(self._ctx, user_id=None, parent=self)
        if dlg.exec():
            self._load_users()

    def _edit_user(self):
        uid = self._get_selected_user_id()
        if uid is None:
            QMessageBox.information(self, "Выбор", "Выберите пользователя.")
            return
        dlg = UserDialog(self._ctx, user_id=uid, parent=self)
        if dlg.exec():
            self._load_users()

    def _change_password(self):
        uid = self._get_selected_user_id()
        if uid is None:
            QMessageBox.information(self, "Выбор", "Выберите пользователя.")
            return
        dlg = ChangePasswordDialog(self._ctx, user_id=uid, parent=self)
        dlg.exec()

    def _toggle_user(self):
        uid = self._get_selected_user_id()
        if uid is None:
            QMessageBox.information(self, "Выбор", "Выберите пользователя.")
            return
        user = self._ctx.db.fetchone(
            "SELECT is_active, full_name FROM users WHERE id=?", (uid,)
        )
        if not user:
            return
        new_state = 0 if user["is_active"] else 1
        action    = "отключить" if user["is_active"] else "включить"
        reply = QMessageBox.question(
            self, "Подтверждение",
            f"{action.capitalize()} пользователя «{user['full_name']}»?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            self._ctx.auth.set_user_active(uid, bool(new_state))
            self._load_users()

    # ── Действия: Доступ ─────────────────────────────────────

    def _grant_access(self, can_edit: bool):
        project_id = self._access_project.currentData()
        if project_id is None:
            return
        for item in self._all_users_list.selectedItems():
            uid = item.data(Qt.ItemDataRole.UserRole)
            self._ctx.auth.grant_access(project_id, uid, can_edit)
        self._load_project_access()

    def _revoke_access(self):
        project_id = self._access_project.currentData()
        if project_id is None:
            return
        for item in self._access_list.selectedItems():
            uid = item.data(Qt.ItemDataRole.UserRole)
            self._ctx.db.execute(
                "DELETE FROM project_access WHERE project_id=? AND user_id=?",
                (project_id, uid)
            )
        self._ctx.db.commit()
        self._load_project_access()

    # ── Экспорт журнала ──────────────────────────────────────

    def _export_audit(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from PySide6.QtWidgets import QFileDialog
        except ImportError:
            QMessageBox.critical(self, "Ошибка", "pip install openpyxl")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить журнал", "audit_log.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "Журнал изменений"
        headers = ["Дата/Время","Пользователь","Проект","Таблица","Действие","Запись №"]
        hdr_font = Font(bold=True, color="F0F0F0")
        hdr_fill = PatternFill("solid", fgColor="1E1E1E")

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(1, ci, h)
            cell.font = hdr_font
            cell.fill = hdr_fill

        for ri in range(self._audit_table.rowCount()):
            for ci in range(self._audit_table.columnCount()):
                item = self._audit_table.item(ri, ci)
                ws.cell(ri + 2, ci + 1, item.text() if item else "")

        wb.save(path)
        QMessageBox.information(self, "Экспорт", f"Журнал сохранён:\n{path}")

    # ── Утилиты ──────────────────────────────────────────────

    def _table_style(self) -> str:
        return f"""
            QTableWidget {{
                background-color: {Colors.BG_BASE};
                color: {Colors.TEXT_PRIMARY};
                gridline-color: {Colors.TABLE_GRID};
                border: 1px solid {Colors.BORDER};
                alternate-background-color: {Colors.TABLE_ROW_ALT};
                font-size: 12px;
            }}
            QTableWidget::item {{ padding: 6px 10px; }}
            QTableWidget::item:selected {{
                background-color: {Colors.TABLE_SELECTED};
            }}
            QHeaderView::section {{
                background-color: {Colors.TABLE_HEADER};
                color: {Colors.TEXT_SECONDARY};
                border: none;
                border-right: 1px solid {Colors.BORDER};
                border-bottom: 1px solid {Colors.BORDER};
                padding: 6px 10px;
                font-size: 11px; font-weight: 600;
            }}
        """

    def _list_style(self) -> str:
        return f"""
            QListWidget {{
                background-color: {Colors.BG_BASE};
                color: {Colors.TEXT_PRIMARY};
                border: 1px solid {Colors.BORDER};
                font-size: 12px;
            }}
            QListWidget::item {{ padding: 8px 10px; border-bottom: 1px solid {Colors.BORDER}; }}
            QListWidget::item:selected {{
                background-color: {Colors.ACCENT_DIM};
                border-left: 3px solid {Colors.ACCENT};
            }}
        """

    @staticmethod
    def _section_label(text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setStyleSheet(
            f"color: {Colors.TEXT_SECONDARY}; font-size: 11px; font-weight: 600;"
        )
        return lbl

    @staticmethod
    def _divider() -> QFrame:
        d = QFrame()
        d.setFixedHeight(1)
        d.setStyleSheet(f"background-color: {Colors.BORDER};")
        return d


# ════════════════════════════════════════════════════════════
# ДИАЛОГИ
# ════════════════════════════════════════════════════════════

class UserDialog(QDialog):
    """Создание / редактирование пользователя."""

    def __init__(self, ctx, user_id: int | None, parent=None):
        super().__init__(parent)
        self._ctx     = ctx
        self._user_id = user_id
        self._user    = None

        if user_id:
            self._user = ctx.db.fetchone(
                "SELECT * FROM users WHERE id=?", (user_id,)
            )

        self.setWindowTitle("Редактировать" if user_id else "Новый пользователь")
        self.setFixedSize(420, 380)
        self.setModal(True)
        self._build_ui()

        if self._user:
            self._fill_form()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(Spacing.XL, Spacing.XL, Spacing.XL, Spacing.XL)
        layout.setSpacing(Spacing.MD)

        form = QFormLayout()
        form.setSpacing(Spacing.SM)

        self._fullname = QLineEdit()
        self._fullname.setFixedHeight(Sizes.INPUT_HEIGHT)
        form.addRow("ФИО *:", self._fullname)

        self._username = QLineEdit()
        self._username.setFixedHeight(Sizes.INPUT_HEIGHT)
        form.addRow("Логин *:", self._username)

        self._department = QLineEdit()
        self._department.setFixedHeight(Sizes.INPUT_HEIGHT)
        form.addRow("Подразделение:", self._department)

        self._role = QComboBox()
        self._role.setFixedHeight(Sizes.INPUT_HEIGHT)
        self._role.addItem("Редактор",       "editor")
        self._role.addItem("Просмотр",       "viewer")
        self._role.addItem("Администратор",  "admin")
        form.addRow("Роль *:", self._role)

        if not self._user_id:
            self._password = QLineEdit()
            self._password.setEchoMode(QLineEdit.EchoMode.Password)
            self._password.setFixedHeight(Sizes.INPUT_HEIGHT)
            self._password.setPlaceholderText("Минимум 8 символов")
            form.addRow("Пароль *:", self._password)

        layout.addLayout(form)

        self._error_lbl = QLabel("")
        self._error_lbl.setStyleSheet(f"color: {Colors.ERROR}; font-size: 12px;")
        self._error_lbl.setVisible(False)
        layout.addWidget(self._error_lbl)

        layout.addStretch()

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save |
            QDialogButtonBox.StandardButton.Cancel
        )
        btns.accepted.connect(self._save)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _fill_form(self):
        self._fullname.setText(self._user.get("full_name", ""))
        self._username.setText(self._user.get("username", ""))
        self._department.setText(self._user.get("department") or "")
        idx = self._role.findData(self._user.get("role", "editor"))
        if idx >= 0:
            self._role.setCurrentIndex(idx)

    def _save(self):
        fullname   = self._fullname.text().strip()
        username   = self._username.text().strip()
        department = self._department.text().strip()
        role       = self._role.currentData()

        if not fullname or not username:
            self._error_lbl.setText("Заполните обязательные поля")
            self._error_lbl.setVisible(True)
            return

        if self._user_id:
            self._ctx.db.execute(
                "UPDATE users SET full_name=?, username=?, department=?, role=? WHERE id=?",
                (fullname, username, department or None, role, self._user_id)
            )
            self._ctx.db.commit()
        else:
            password = self._password.text()
            if len(password) < 8:
                self._error_lbl.setText("Пароль — минимум 8 символов")
                self._error_lbl.setVisible(True)
                return
            ok, msg = self._ctx.auth.create_user(
                username, password, fullname, department, role
            )
            if not ok:
                self._error_lbl.setText(msg)
                self._error_lbl.setVisible(True)
                return

        self.accept()


class ChangePasswordDialog(QDialog):
    """Смена пароля пользователя."""

    def __init__(self, ctx, user_id: int, parent=None):
        super().__init__(parent)
        self._ctx     = ctx
        self._user_id = user_id

        user = ctx.db.fetchone("SELECT full_name FROM users WHERE id=?", (user_id,))
        name = user["full_name"] if user else ""

        self.setWindowTitle(f"Сменить пароль — {name}")
        self.setFixedSize(360, 240)
        self.setModal(True)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(Spacing.XL, Spacing.XL, Spacing.XL, Spacing.XL)
        layout.setSpacing(Spacing.MD)

        form = QFormLayout()

        self._pwd1 = QLineEdit()
        self._pwd1.setEchoMode(QLineEdit.EchoMode.Password)
        self._pwd1.setFixedHeight(Sizes.INPUT_HEIGHT)
        self._pwd1.setPlaceholderText("Минимум 8 символов")
        form.addRow("Новый пароль:", self._pwd1)

        self._pwd2 = QLineEdit()
        self._pwd2.setEchoMode(QLineEdit.EchoMode.Password)
        self._pwd2.setFixedHeight(Sizes.INPUT_HEIGHT)
        form.addRow("Повторите:", self._pwd2)

        layout.addLayout(form)

        self._error_lbl = QLabel("")
        self._error_lbl.setStyleSheet(f"color: {Colors.ERROR}; font-size: 12px;")
        self._error_lbl.setVisible(False)
        layout.addWidget(self._error_lbl)

        layout.addStretch()

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save |
            QDialogButtonBox.StandardButton.Cancel
        )
        btns.accepted.connect(self._save)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _save(self):
        p1 = self._pwd1.text()
        p2 = self._pwd2.text()

        if len(p1) < 8:
            self._error_lbl.setText("Минимум 8 символов")
            self._error_lbl.setVisible(True)
            return
        if p1 != p2:
            self._error_lbl.setText("Пароли не совпадают")
            self._error_lbl.setVisible(True)
            self._pwd2.clear()
            return

        ok, msg = self._ctx.auth.change_password(self._user_id, p1)
        if ok:
            QMessageBox.information(self, "Готово", "Пароль изменён.")
            self.accept()
        else:
            self._error_lbl.setText(msg)
            self._error_lbl.setVisible(True)