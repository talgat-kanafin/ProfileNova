"""
ui/screens/station_page.py
===========================
Раздел "Станция" — соответствует вкладке tshSingleLine старой программы.

Навигация: Станция → Путь → данные
Вкладки:
  1. Основные данные  (tbStationWayData)
  2. Устройства       (tbStationDeviceLocation)
"""

import os, sys
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QTabWidget, QPushButton, QComboBox, QFrame,
    QMessageBox, QLineEdit, QApplication
)
from PySide6.QtCore import Qt, QThread, Signal

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from ui.theme import Colors, Typography, Spacing, Sizes
from ui.screens.mainline_page import RailwayTable, DataLoaderThread


# ── Определения колонок ──────────────────────────────────────

STATION_WAY_DATA_COLS = [
    ("RecordNumber",   "№"),
    ("Km",             "Км"),
    ("Picket",         "Пк"),
    ("Meter",          "Метр"),
    ("PicketLength",   "Длина пк"),
    ("Ballast",        "Балласт"),
    ("Rail",           "Рельс"),
    ("SUklon",         "Уклон"),
    ("LowPurityBallast","Загр.балл."),
]

STATION_DEVICE_COLS = [
    ("RecordNumber", "№"),
    ("Kilometer",    "Км"),
    ("Picket",       "Пк"),
    ("Meter",        "Метр"),
    ("DeviceId",     "Устройство"),
    ("Note",         "Примечание"),
]


class StationPage(QWidget):
    """Страница раздела Станция."""

    def __init__(self, ctx, project_id: int, is_editing: bool, parent=None):
        super().__init__(parent)
        self._ctx        = ctx
        self._project_id = project_id
        self._is_editing = is_editing
        self._station_id = None
        self._way_id     = None
        self._loaders    = []

        self._build_ui()
        self._load_stations()

    # ── UI ───────────────────────────────────────────────────

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # Тулбар — выбор станции и пути
        tb = QFrame()
        tb.setFixedHeight(Sizes.TOOLBAR_HEIGHT)
        tb.setStyleSheet(f"""
            background-color: {Colors.BG_SURFACE};
            border-bottom: 1px solid {Colors.BORDER};
        """)
        tb_layout = QHBoxLayout(tb)
        tb_layout.setContentsMargins(Spacing.LG, 0, Spacing.LG, 0)
        tb_layout.setSpacing(Spacing.MD)

        tb_layout.addWidget(QLabel("Станция:"))
        self._station_combo = QComboBox()
        self._station_combo.setFixedWidth(260)
        self._station_combo.setFixedHeight(32)
        self._station_combo.currentIndexChanged.connect(self._on_station_changed)
        tb_layout.addWidget(self._station_combo)

        btn_add_st = QPushButton("＋")
        btn_add_st.setFixedSize(32, 32)
        btn_add_st.setToolTip("Добавить станцию")
        btn_add_st.clicked.connect(self._add_station)
        tb_layout.addWidget(btn_add_st)

        tb_layout.addSpacing(Spacing.MD)
        tb_layout.addWidget(QLabel("Путь:"))
        self._way_combo = QComboBox()
        self._way_combo.setFixedWidth(220)
        self._way_combo.setFixedHeight(32)
        self._way_combo.currentIndexChanged.connect(self._on_way_changed)
        tb_layout.addWidget(self._way_combo)
        btn_add_way = QPushButton("+")
        btn_add_way.setFixedSize(32, 32)
        btn_add_way.setFont(Typography.get_font(16, bold=True))
        btn_add_way.setToolTip("Добавить путь")
        btn_add_way.setStyleSheet(f"""
            QPushButton {{
                background-color: {Colors.BG_ELEVATED};
                color: {Colors.TEXT_PRIMARY};
                border: 1px solid {Colors.BORDER};
                border-radius: 4px;
            }}
            QPushButton:hover {{
                background-color: {Colors.ACCENT};
                color: white;
                border-color: {Colors.ACCENT};
            }}
        """)
        btn_add_way.clicked.connect(self._add_way)
        tb_layout.addWidget(btn_add_way)

        tb_layout.addStretch()

        # Поиск
        self._search = QLineEdit()
        self._search.setPlaceholderText("Поиск по таблице...")
        self._search.setFixedWidth(200)
        self._search.setFixedHeight(32)
        self._search.textChanged.connect(self._filter_current_table)
        tb_layout.addWidget(self._search)

        # Экспорт
        btn_export = QPushButton("📊 Excel")
        btn_export.setFixedHeight(32)
        btn_export.clicked.connect(self.export_excel)
        tb_layout.addWidget(btn_export)

        root.addWidget(tb)

        btn_save = QPushButton("💾 Сохранить")
        btn_save.setFixedHeight(32)
        btn_save.setStyleSheet(f"""
            QPushButton {{
                background-color: {Colors.ACCENT}; color: white;
                border: none; border-radius: 4px;
                font-weight: 600; padding: 0 12px;
            }}
            QPushButton:hover {{ background-color: {Colors.ACCENT_HOVER}; }}
        """)
        btn_save.clicked.connect(self._save_changes)
        tb_layout.addWidget(btn_save)

        # Счётчик
        self._count_lbl = QLabel("")
        self._count_lbl.setStyleSheet(f"""
            background-color: {Colors.BG_SURFACE};
            color: {Colors.TEXT_SECONDARY};
            padding: 3px {Spacing.LG}px;
            font-size: 11px;
            border-bottom: 1px solid {Colors.BORDER};
        """)
        root.addWidget(self._count_lbl)

        # Вкладки
        self._tabs = QTabWidget()
        self._tabs.setStyleSheet(f"""
            QTabWidget::pane {{
                border: none;
                background-color: {Colors.BG_BASE};
            }}
            QTabBar::tab {{
                background-color: {Colors.BG_SURFACE};
                color: {Colors.TEXT_SECONDARY};
                padding: 8px 18px;
                border: none;
                border-bottom: 2px solid transparent;
                font-size: 12px;
            }}
            QTabBar::tab:selected {{
                color: {Colors.TEXT_PRIMARY};
                border-bottom: 2px solid {Colors.ACCENT};
            }}
            QTabBar::tab:hover {{
                color: {Colors.TEXT_PRIMARY};
                background-color: {Colors.BG_ELEVATED};
            }}
        """)

        self._tbl_waydata = RailwayTable(self._is_editing)
        self._tbl_waydata.cellChanged.connect(self._on_cell_changed)
        self._tbl_devices = RailwayTable(self._is_editing)

        self._tabs.addTab(self._tbl_waydata, "Основные данные")
        self._tabs.addTab(self._tbl_devices, "Устройства")
        self._tabs.currentChanged.connect(lambda _: self._search.clear())

        root.addWidget(self._tabs, 1)

    # ── Загрузка ─────────────────────────────────────────────

    def _load_stations(self):
        rows = self._ctx.db.fetchall(
            "SELECT StationId, Name FROM tbStation "
            "WHERE project_id=? ORDER BY StationId",
            (self._project_id,)
        )
        self._station_combo.blockSignals(True)
        self._station_combo.clear()
        for r in rows:
            name = r.get("Name") or f"Станция {r['StationId']}"
            self._station_combo.addItem(name, r["StationId"])
        self._station_combo.blockSignals(False)

        if self._station_combo.count() > 0:
            self._station_id = self._station_combo.currentData()
            self._load_ways_for_station()

    def _on_station_changed(self):
        self._station_id = self._station_combo.currentData()
        if self._station_id is not None:
            self._load_ways_for_station()

    def _load_ways_for_station(self):
        rows = self._ctx.db.fetchall(
            "SELECT WayId, Name, stKm, enKm FROM tbStationWay "
            "WHERE project_id=? AND StationId=? ORDER BY WayId",
            (self._project_id, self._station_id)
        )
        self._way_combo.blockSignals(True)
        self._way_combo.clear()
        for r in rows:
            name = r.get("Name") or f"Путь {r['WayId']}"
            km   = ""
            if r.get("stKm") and r.get("enKm"):
                km = f"  ({r['stKm']} — {r['enKm']} км)"
            self._way_combo.addItem(name + km, r["WayId"])
        self._way_combo.blockSignals(False)

        if self._way_combo.count() > 0:
            self._way_id = self._way_combo.currentData()
            self._load_data()
        else:
            self._tbl_waydata.setRowCount(0)
            self._tbl_devices.setRowCount(0)
            self._count_lbl.setText("Нет путей для выбранной станции")

    def _on_way_changed(self):
        self._way_id = self._way_combo.currentData()
        if self._way_id is not None:
            self._load_data()

    def _load_data(self):
        if not self._station_id or not self._way_id:
            return

        pid = self._project_id
        sid = self._station_id
        wid = self._way_id

        # Основные данные
        self._load_table(
            self._tbl_waydata, STATION_WAY_DATA_COLS,
            "SELECT * FROM tbStationWayData "
            "WHERE project_id=? AND StationId=? AND StationWayId=? "
            "ORDER BY Km, Picket",
            (pid, sid, wid)
        )

        # Устройства
        self._load_table(
            self._tbl_devices, STATION_DEVICE_COLS,
            "SELECT d.RecordNumber, d.Kilometer, d.Picket, d.Meter, "
            "       dev.DeviceName as DeviceId, d.Note "
            "FROM tbStationDeviceLocation d "
            "LEFT JOIN tbDevice dev ON dev.DeviceId = d.DeviceId "
            "WHERE d.project_id=? AND d.StationId=? AND d.WayId=? "
            "ORDER BY d.Kilometer, d.Picket",
            (pid, sid, wid)
        )

    def _load_table(self, table_widget: RailwayTable,
                    col_defs: list, sql: str, params: tuple):
        loader = DataLoaderThread(self._ctx.db, sql, params)
        loader.finished.connect(
            lambda cols, rows, tw=table_widget, cd=col_defs:
                self._on_loaded(tw, cd, cols, rows)
        )
        loader.error.connect(lambda e: print(f"Station load error: {e}"))
        loader.start()
        self._loaders.append(loader)

    def _on_loaded(self, table, col_defs, db_cols, rows):
        try:
            table.cellChanged.disconnect()
        except Exception:
            pass
        table.fill(col_defs, db_cols, rows)
        if table is self._tbl_waydata:
            table.cellChanged.connect(self._on_cell_changed)
            self._count_lbl.setText(...)

    # ── Фильтрация ───────────────────────────────────────────

    def _filter_current_table(self, text: str):
        idx = self._tabs.currentIndex()
        tbl = self._tbl_waydata if idx == 0 else self._tbl_devices
        text = text.lower()
        for row in range(tbl.rowCount()):
            match = not text
            if not match:
                for col in range(tbl.columnCount()):
                    item = tbl.item(row, col)
                    if item and text in item.text().lower():
                        match = True
                        break
            tbl.setRowHidden(row, not match)

    # ── Экспорт ──────────────────────────────────────────────

    def export_excel(self):
        if not self._station_id or not self._way_id:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from PySide6.QtWidgets import QFileDialog
        except ImportError:
            QMessageBox.critical(self, "Ошибка", "pip install openpyxl")
            return

        station = self._station_combo.currentText()
        way     = self._way_combo.currentText().split("  ")[0]
        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить Excel",
            f"Станция_{station}_{way}.xlsx",
            "Excel (*.xlsx)"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        hdr_font = Font(bold=True, color="F0F0F0")
        hdr_fill = PatternFill("solid", fgColor="1E1E1E")

        for tbl, title, cols in [
            (self._tbl_waydata, "Основные данные", STATION_WAY_DATA_COLS),
            (self._tbl_devices, "Устройства",      STATION_DEVICE_COLS),
        ]:
            ws = wb.active if title == "Основные данные" else wb.create_sheet(title)
            ws.title = title
            for ci, (_, h) in enumerate(cols, 1):
                cell = ws.cell(1, ci, h)
                cell.font = hdr_font
                cell.fill = hdr_fill
            for ri in range(tbl.rowCount()):
                for ci in range(tbl.columnCount()):
                    item = tbl.item(ri, ci)
                    ws.cell(ri + 2, ci + 1, item.text() if item else "")

        wb.save(path)
        QMessageBox.information(self, "Экспорт завершён", f"Файл сохранён:\n{path}")

    def _add_station(self):
        from ui.dialogs.add_station_dialog import AddStationDialog
        dlg = AddStationDialog(self._ctx, self._project_id, self)
        if dlg.exec():
            self._load_stations()

    def _save_changes(self):
        if not self._is_editing:
            return
        from PySide6.QtWidgets import QMessageBox
        from ui.screens.mainline_page import STATION_WAY_DATA_COLS
    
        saved = errors = 0
        col_defs = STATION_WAY_DATA_COLS
    
        for row in range(self._tbl_waydata.rowCount()):
            rec_item = self._tbl_waydata.item(row, 0)
            if not rec_item or not rec_item.text():
                continue
            try:
                record_num = int(rec_item.text())
            except ValueError:
                continue
    
            updates = {}
            for col_idx, (field, _) in enumerate(col_defs):
                if field == "RecordNumber":
                    continue
                item = self._tbl_waydata.item(row, col_idx)
                if item:
                    updates[field] = item.text().strip() or None
    
            if not updates:
                continue
    
            sets   = ", ".join(f"{k}=?" for k in updates.keys())
            values = list(updates.values()) + [record_num, self._project_id]
            try:
                self._ctx.db.execute(
                    f"UPDATE tbStationWayData SET {sets} "
                    f"WHERE RecordNumber=? AND project_id=?",
                    values
                )
                saved += 1
            except Exception as e:
                errors += 1
    
        self._ctx.db.commit()
        QMessageBox.information(self, "Сохранено",
            f"Сохранено строк: {saved}\nОшибок: {errors}")
        
    def _on_cell_changed(self, row: int, col: int):
        from PySide6.QtGui import QColor
        item = self._tbl_waydata.item(row, col)
        if item:
            item.setBackground(QColor("#FF6B0033"))

    def _add_way(self):
        if not self._station_id:
            return
        from ui.dialogs.add_way_dialog import AddWayDialog
        dlg = AddWayDialog(self._ctx, self._project_id, self)
        if dlg.exec():
            self._load_ways_for_station()