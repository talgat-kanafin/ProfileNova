"""
ui/screens/mainline_page.py
============================
Раздел "Перегон" — главный раздел путевых данных.

Вкладки:
  1. Данные профиля  (tbWayData)
  2. Кривые          (tbLeftCurve / tbRightCurve / tbCurve)
  3. Устройства      (tbDeviceLocation)
  4. Стыки           (tbLeftIsolatedJoint / tbRightIsolatedJoint)
  5. Плети/Вставки   (tbLeftPleti/tbRightPleti/tbLeftVstavki/tbRightVstavki)
  6. Выправка        (tbStraighteningData)
"""

import os, sys
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QTabWidget, QTableWidget, QTableWidgetItem,
    QPushButton, QComboBox, QFrame, QHeaderView,
    QAbstractItemView, QMessageBox, QApplication,
    QLineEdit, QMenu
)
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui  import QKeySequence, QShortcut

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from ui.theme import Colors, Typography, Spacing, Sizes


# ── Заголовки столбцов ───────────────────────────────────────

WAY_DATA_COLS = [
    ("RecordNumber", "№"),
    ("Km",           "Км"),
    ("Picket",       "Пк"),
    ("Meter",        "Метр"),
    ("PicketLength", "Длина пк"),
    ("Left_Ballast", "Балласт Л"),
    ("Left_Rail",    "Рельс Л"),
    ("Left_SUklon",  "Уклон Л"),
    ("TrackSpacing", "Межпутье"),
    ("Right_Ballast","Балласт П"),
    ("Right_Rail",   "Рельс П"),
    ("Right_SUklon", "Уклон П"),
    ("Ground",       "Земля"),
    ("LowPurityBallast","Загр.балл."),
    ("GroundEdge",   "Бровка"),
    ("Left_GEToAxisDistance", "Расст.Л"),
    ("Right_GEToAxisDistance","Расст.П"),
]

CURVE_COLS = [
    ("RecordNumber","№"),
    ("AngleDegree", "Угол°"),
    ("AngleMinute", "Угол'"),
    ("Radius",      "Радиус"),
    ("Tangent",     "Тангенс"),
    ("Curvature",   "Кривизна"),
    ("Length",      "Длина"),
    ("startKm",     "От Км"),
    ("startPk",     "От Пк"),
    ("startM",      "От М"),
    ("endKm",       "До Км"),
    ("endPk",       "До Пк"),
    ("endM",        "До М"),
    ("Elevation",   "Возв."),
]

DEVICE_COLS = [
    ("RecordNumber","№"),
    ("Kilometer",   "Км"),
    ("Picket",      "Пк"),
    ("Meter",       "Метр"),
    ("DeviceId",    "Устройство"),
    ("LeftWay",     "Лев.путь"),
    ("RightWay",    "Пр.путь"),
    ("Left_CurrentDistanceToAxis",  "Расст.Л"),
    ("Right_CurrentDistanceToAxis", "Расст.П"),
    ("Note",        "Примечание"),
]

JOINT_COLS = [
    ("RecordNumber","№"),
    ("Km",          "Км"),
    ("Picket",      "Пк"),
    ("Meter",       "Метр"),
    ("CurrentKm",   "Тек.Км"),
    ("CurrentPk",   "Тек.Пк"),
    ("CurrentM",    "Тек.М"),
    ("ProjectKm",   "Пр.Км"),
    ("ProjectPk",   "Пр.Пк"),
    ("ProjectM",    "Пр.М"),
]

PLETI_COLS = [
    ("RecordNumber","№"),
    ("RStartKm","R-Нач.Км"),("RStartPk","R-Нач.Пк"),("RStartM","R-Нач.М"),
    ("REndKm",  "R-Кон.Км"),("REndPk",  "R-Кон.Пк"),("REndM",  "R-Кон.М"),
    ("LStartKm","L-Нач.Км"),("LStartPk","L-Нач.Пк"),("LStartM","L-Нач.М"),
    ("LEndKm",  "L-Кон.Км"),("LEndPk",  "L-Кон.Пк"),("LEndM",  "L-Кон.М"),
]

VSTAVKI_COLS = [
    ("RecordNumber","№"),
    ("StartKm","Нач.Км"),("StartPk","Нач.Пк"),("StartM","Нач.М"),
    ("EndKm",  "Кон.Км"),("EndPk",  "Кон.Пк"),("EndM",  "Кон.М"),
    ("Vstavka","Вставка"),
]

STRAIGHT_COLS = [
    ("RecordNumber",        "№"),
    ("Km",                  "Км"),
    ("Picket",              "Пк"),
    ("Meter",               "Метр"),
    ("Left_Straightening",  "Выправка Л"),
    ("CurrentTrackSpacing", "Межпутье"),
    ("Right_Straightening", "Выправка П"),
]


class DataLoaderThread(QThread):
    finished = Signal(list, list)  # columns, rows
    error    = Signal(str)

    def __init__(self, db, sql, params=()):
        super().__init__()
        self._db     = db
        self._sql    = sql
        self._params = params

    def run(self):
        try:
            rows = self._db.fetchall(self._sql, self._params)
            if rows:
                cols = list(rows[0].keys())
                data = [list(r.values()) for r in rows]
            else:
                cols, data = [], []
            self.finished.emit(cols, data)
        except Exception as e:
            self.error.emit(str(e))


class RailwayTable(QTableWidget):
    """
    Базовая таблица с поддержкой:
    - Ctrl+C / Ctrl+V / Ctrl+X / Ctrl+Z
    - Контекстное меню
    - Фильтрация
    - Только для чтения если не редактирование
    """

    def __init__(self, is_editing: bool, parent=None):
        super().__init__(parent)
        self._is_editing  = is_editing
        self._undo_stack  = []

        self.setAlternatingRowColors(True)
        self.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        self.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked
            if is_editing else
            QAbstractItemView.EditTrigger.NoEditTriggers
        )
        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.horizontalHeader().setStretchLastSection(True)
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self._context_menu)
        self.setStyleSheet(f"""
            QTableWidget {{
                background-color: {Colors.BG_BASE};
                color: {Colors.TEXT_PRIMARY};
                gridline-color: {Colors.TABLE_GRID};
                border: none;
                font-size: 12px;
            }}
            QTableWidget::item {{ padding: 3px 6px; }}
            QTableWidget::item:selected {{
                background-color: {Colors.TABLE_SELECTED};
                color: {Colors.TEXT_PRIMARY};
            }}
            QHeaderView::section {{
                background-color: {Colors.TABLE_HEADER};
                color: {Colors.TEXT_SECONDARY};
                border: none;
                border-right: 1px solid {Colors.BORDER};
                border-bottom: 1px solid {Colors.BORDER};
                padding: 5px 6px;
                font-size: 11px;
                font-weight: 600;
            }}
        """)

        # Шорткаты
        QShortcut(QKeySequence.StandardKey.Copy,      self, self.copy_selection)
        QShortcut(QKeySequence.StandardKey.Paste,     self, self.paste_selection)
        QShortcut(QKeySequence.StandardKey.Cut,       self, self.cut_selection)
        QShortcut(QKeySequence.StandardKey.Undo,      self, self.undo)
        QShortcut(QKeySequence.StandardKey.SelectAll, self, self.selectAll)

    def fill(self, col_defs: list, db_cols: list, rows: list):
        """
        col_defs = [(db_field, display_name), ...]
        db_cols  = список колонок из БД
        rows     = список строк (list of lists)
        """
        self.setSortingEnabled(False)
        self.setRowCount(0)
        self.setColumnCount(len(col_defs))
        self.setHorizontalHeaderLabels([d for _, d in col_defs])

        # Маппинг: db_field → индекс в db_cols
        col_map = {c: i for i, c in enumerate(db_cols)}

        for row_data in rows:
            row_idx = self.rowCount()
            self.insertRow(row_idx)
            for col_idx, (field, _) in enumerate(col_defs):
                db_idx = col_map.get(field, -1)
                val    = str(row_data[db_idx]) if db_idx >= 0 and row_data[db_idx] is not None else ""
                item   = QTableWidgetItem(val)
                item.setFlags(
                    Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled |
                    (Qt.ItemFlag.ItemIsEditable if self._is_editing else Qt.ItemFlag.NoItemFlags)
                )
                self.setItem(row_idx, col_idx, item)

        self.setSortingEnabled(True)

    # ── Clipboard ────────────────────────────────────────────

    def copy_selection(self):
        selected = self.selectedRanges()
        if not selected:
            return
        r = selected[0]
        lines = []
        for row in range(r.topRow(), r.bottomRow() + 1):
            cells = []
            for col in range(r.leftColumn(), r.rightColumn() + 1):
                item = self.item(row, col)
                cells.append(item.text() if item else "")
            lines.append("\t".join(cells))
        QApplication.clipboard().setText("\n".join(lines))

    def paste_selection(self):
        if not self._is_editing:
            return
        text = QApplication.clipboard().text()
        if not text:
            return
        self._save_undo()
        start = self.currentIndex()
        if not start.isValid():
            return
        r0, c0 = start.row(), start.column()
        for r, line in enumerate(text.splitlines()):
            for c, val in enumerate(line.split("\t")):
                row, col = r0 + r, c0 + c
                if row >= self.rowCount():
                    self.insertRow(row)
                if col < self.columnCount():
                    item = QTableWidgetItem(val)
                    self.setItem(row, col, item)

    def cut_selection(self):
        if not self._is_editing:
            return
        self.copy_selection()
        self._save_undo()
        for idx in self.selectedIndexes():
            item = self.item(idx.row(), idx.column())
            if item:
                item.setText("")

    def undo(self):
        if not self._undo_stack or not self._is_editing:
            return
        state = self._undo_stack.pop()
        for (r, c), text in state.items():
            item = self.item(r, c)
            if item:
                item.setText(text)

    def _save_undo(self):
        state = {}
        for idx in self.selectedIndexes():
            item = self.item(idx.row(), idx.column())
            state[(idx.row(), idx.column())] = item.text() if item else ""
        if state:
            self._undo_stack.append(state)
            if len(self._undo_stack) > 30:
                self._undo_stack.pop(0)

    # ── Контекстное меню ─────────────────────────────────────

    def _context_menu(self, pos):
        menu = QMenu(self)
        menu.setStyleSheet(f"""
            QMenu {{
                background-color: {Colors.BG_OVERLAY};
                color: {Colors.TEXT_PRIMARY};
                border: 1px solid {Colors.BORDER};
                padding: 4px 0;
            }}
            QMenu::item {{ padding: 7px 20px 7px 12px; }}
            QMenu::item:selected {{ background-color: {Colors.BG_ELEVATED}; }}
            QMenu::separator {{ height:1px; background:{Colors.BORDER}; margin:3px 0; }}
        """)

        act_copy = menu.addAction("Копировать  Ctrl+C")
        act_copy.triggered.connect(self.copy_selection)

        menu.addSeparator()
        act_split = menu.addAction("✂️ Разделить текст по разделителям")
        act_split.triggered.connect(self._split_text)

        if self._is_editing:
            act_paste = menu.addAction("Вставить  Ctrl+V")
            act_paste.triggered.connect(self.paste_selection)

            act_cut = menu.addAction("Вырезать  Ctrl+X")
            act_cut.triggered.connect(self.cut_selection)

            menu.addSeparator()

            act_clear = menu.addAction("Очистить ячейки")
            act_clear.triggered.connect(self._clear_selected)

        menu.addSeparator()
        act_select_all = menu.addAction("Выделить всё  Ctrl+A")
        act_select_all.triggered.connect(self.selectAll)

        menu.exec(self.viewport().mapToGlobal(pos))

    def _clear_selected(self):
        self._save_undo()
        for idx in self.selectedIndexes():
            item = self.item(idx.row(), idx.column())
            if item:
                item.setText("")

    def _split_text(self):
        selected = self.selectedIndexes()
        if not selected:
            return
    
        # Берём значения из первого выделенного столбца
        col = selected[0].column()
        rows_idx = sorted(set(idx.row() for idx in selected))
        values = []
        for row in rows_idx:
            item = self.item(row, col)
            values.append(item.text() if item else "")
    
        from ui.dialogs.split_text_dialog import SplitTextDialog
        dlg = SplitTextDialog(values, self)
        if dlg.exec() != dlg.DialogCode.Accepted:
            return
    
        results = dlg.get_results()
        # Записать результаты начиная с выделенного столбца
        for col_offset, col_values in results.items():
            target_col = col + col_offset
            if target_col >= self.columnCount():
                break
            for i, row in enumerate(rows_idx):
                if i < len(col_values):
                    item = self.item(row, target_col)
                    if item is None:
                        item = QTableWidgetItem()
                        self.setItem(row, target_col, item)
                    item.setText(col_values[i])


class MainlinePage(QWidget):
    """Страница раздела Перегон."""

    def __init__(self, ctx, project_id: int, is_editing: bool, parent=None):
        super().__init__(parent)
        self._ctx        = ctx
        self._project_id = project_id
        self._is_editing = is_editing
        self._way_id     = None
        self._loaders    = []

        self._build_ui()
        self._load_ways()

    # ── UI ───────────────────────────────────────────────────

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # Тулбар
        tb = QFrame()
        tb.setFixedHeight(Sizes.TOOLBAR_HEIGHT)
        tb.setStyleSheet(f"""
            background-color: {Colors.BG_SURFACE};
            border-bottom: 1px solid {Colors.BORDER};
        """)
        tb_layout = QHBoxLayout(tb)
        tb_layout.setContentsMargins(Spacing.LG, 0, Spacing.LG, 0)
        tb_layout.setSpacing(Spacing.MD)

        tb_layout.addWidget(QLabel("Путь:"))

        self._way_combo = QComboBox()
        self._way_combo.setFixedWidth(280)
        self._way_combo.setFixedHeight(32)
        self._way_combo.currentIndexChanged.connect(self._on_way_changed)
        tb_layout.addWidget(self._way_combo)

        btn_add_way = QPushButton("+")
        btn_add_way.setFixedSize(32, 32)
        btn_add_way.setFont(Typography.get_font(16, bold=True))
        btn_add_way.setToolTip("Добавить новый путь")
        btn_add_way.setStyleSheet(f"""
            QPushButton {{
                background-color: {Colors.BG_ELEVATED};
                color: {Colors.TEXT_PRIMARY};
                border: 1px solid {Colors.BORDER};
                border-radius: 4px;
                font-size: 16px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: {Colors.ACCENT};
                color: white;
                border-color: {Colors.ACCENT};
            }}
        """)
        btn_add_way.clicked.connect(self._add_way)
        tb_layout.addWidget(btn_add_way)

        tb_layout.addStretch()

        # Поиск
        self._search = QLineEdit()
        self._search.setPlaceholderText("Поиск по таблице...")
        self._search.setFixedWidth(200)
        self._search.setFixedHeight(32)
        self._search.textChanged.connect(self._filter_current_table)
        tb_layout.addWidget(self._search)

        # Кнопки
        btn_graph = QPushButton("📈 График")
        btn_graph.setFixedHeight(32)
        btn_graph.clicked.connect(self._show_graph)
        tb_layout.addWidget(btn_graph)

        btn_export = QPushButton("📊 Excel")
        btn_export.setFixedHeight(32)
        btn_export.clicked.connect(self.export_excel)
        tb_layout.addWidget(btn_export)

        root.addWidget(tb)

        btn_add_row = QPushButton("＋ Строка")
        btn_add_row.setFixedHeight(32)
        btn_add_row.clicked.connect(self._add_row)
        tb_layout.addWidget(btn_add_row)

        btn_save = QPushButton("💾 Сохранить")
        btn_save.setFixedHeight(32)
        btn_save.setStyleSheet(f"""
            QPushButton {{
                background-color: {Colors.ACCENT};
                color: white; border: none; border-radius: 4px;
                font-weight: 600; padding: 0 12px;
            }}
            QPushButton:hover {{ background-color: {Colors.ACCENT_HOVER}; }}
        """)
        btn_save.clicked.connect(self._save_changes)
        tb_layout.addWidget(btn_save)

        # Счётчик строк
        self._count_lbl = QLabel("")
        self._count_lbl.setStyleSheet(f"""
            background-color: {Colors.BG_SURFACE};
            color: {Colors.TEXT_SECONDARY};
            padding: 3px {Spacing.LG}px;
            font-size: 11px;
            border-bottom: 1px solid {Colors.BORDER};
        """)
        root.addWidget(self._count_lbl)

        # Вкладки
        self._tabs = QTabWidget()
        self._tabs.setStyleSheet(f"""
            QTabWidget::pane {{
                border: none;
                background-color: {Colors.BG_BASE};
            }}
            QTabBar::tab {{
                background-color: {Colors.BG_SURFACE};
                color: {Colors.TEXT_SECONDARY};
                padding: 8px 18px;
                border: none;
                border-bottom: 2px solid transparent;
                font-size: 12px;
            }}
            QTabBar::tab:selected {{
                color: {Colors.TEXT_PRIMARY};
                border-bottom: 2px solid {Colors.ACCENT};
            }}
            QTabBar::tab:hover {{
                color: {Colors.TEXT_PRIMARY};
                background-color: {Colors.BG_ELEVATED};
            }}
        """)
        self._tabs.currentChanged.connect(self._on_tab_changed)

        # Создать таблицы для каждой вкладки
        self._tbl_waydata   = RailwayTable(self._is_editing)
        self._tbl_curves_l  = RailwayTable(self._is_editing)
        self._tbl_curves_r  = RailwayTable(self._is_editing)
        self._tbl_curves_m  = RailwayTable(self._is_editing)
        self._tbl_devices   = RailwayTable(self._is_editing)
        self._tbl_joints_l  = RailwayTable(self._is_editing)
        self._tbl_joints_r  = RailwayTable(self._is_editing)
        self._tbl_pleti_l   = RailwayTable(self._is_editing)
        self._tbl_pleti_r   = RailwayTable(self._is_editing)
        self._tbl_vstavki_l = RailwayTable(self._is_editing)
        self._tbl_vstavki_r = RailwayTable(self._is_editing)
        self._tbl_straight  = RailwayTable(self._is_editing)

        self._tbl_waydata.cellChanged.connect(self._on_cell_changed)

        # Вкладка 1: Данные профиля
        self._tabs.addTab(self._tbl_waydata, "Данные профиля")

        # Вкладка 2: Кривые
        curves_widget = QTabWidget()
        curves_widget.setStyleSheet(self._tabs.styleSheet())
        curves_widget.addTab(self._tbl_curves_l, "Левые")
        curves_widget.addTab(self._tbl_curves_r, "Правые")
        curves_widget.addTab(self._tbl_curves_m, "Главный путь")
        self._tabs.addTab(curves_widget, "Кривые")

        # Вкладка 3: Устройства
        self._tabs.addTab(self._tbl_devices, "Устройства")

        # Вкладка 4: Стыки
        joints_widget = QTabWidget()
        joints_widget.setStyleSheet(self._tabs.styleSheet())
        joints_widget.addTab(self._tbl_joints_l, "Левые")
        joints_widget.addTab(self._tbl_joints_r, "Правые")
        self._tabs.addTab(joints_widget, "Стыки")

        # Вкладка 5: Плети и вставки
        pleti_widget = QTabWidget()
        pleti_widget.setStyleSheet(self._tabs.styleSheet())
        pleti_widget.addTab(self._tbl_pleti_l,   "Плети Л")
        pleti_widget.addTab(self._tbl_pleti_r,   "Плети П")
        pleti_widget.addTab(self._tbl_vstavki_l, "Вставки Л")
        pleti_widget.addTab(self._tbl_vstavki_r, "Вставки П")
        self._tabs.addTab(pleti_widget, "Плети / Вставки")

        # Вкладка 6: Выправка
        self._tabs.addTab(self._tbl_straight, "Выправка")

        root.addWidget(self._tabs, 1)

    # ── Загрузка данных ──────────────────────────────────────

    def _load_ways(self):
        rows = self._ctx.db.fetchall(
            "SELECT WayId, Name, stKm, enKm FROM tbWayInfo WHERE project_id=? ORDER BY WayId",
            (self._project_id,)
        )
        self._way_combo.blockSignals(True)
        self._way_combo.clear()
        for r in rows:
            name = r.get("Name") or f"Путь {r['WayId']}"
            km   = f"  ({r.get('stKm','?')} — {r.get('enKm','?')} км)"
            self._way_combo.addItem(name + km, r["WayId"])
        self._way_combo.blockSignals(False)

        if self._way_combo.count() > 0:
            self._way_id = self._way_combo.currentData()
            self._load_all_tabs()

    def _on_way_changed(self):
        self._way_id = self._way_combo.currentData()
        if self._way_id is not None:
            self._load_all_tabs()

    def _load_all_tabs(self):
        pid = self._project_id
        wid = self._way_id
        self._load_table(self._tbl_waydata,   WAY_DATA_COLS,
            "SELECT * FROM tbWayData WHERE project_id=? AND WayId=? ORDER BY Km, Picket",
            (pid, wid))
        self._load_table(self._tbl_curves_l,  CURVE_COLS,
            "SELECT * FROM tbLeftCurve WHERE project_id=? AND WayId=? ORDER BY RecordNumber",
            (pid, wid))
        self._load_table(self._tbl_curves_r,  CURVE_COLS,
            "SELECT * FROM tbRightCurve WHERE project_id=? AND WayId=? ORDER BY RecordNumber",
            (pid, wid))
        self._load_table(self._tbl_curves_m,  CURVE_COLS[:-1],  # без Elevation
            "SELECT * FROM tbCurve WHERE project_id=? AND WayId=? ORDER BY RecordNumber",
            (pid, wid))
        self._load_table(self._tbl_devices,   DEVICE_COLS,
            "SELECT * FROM tbDeviceLocation WHERE project_id=? AND WayId=? ORDER BY Kilometer, Picket",
            (pid, wid))
        self._load_table(self._tbl_joints_l,  JOINT_COLS,
            "SELECT * FROM tbLeftIsolatedJoint WHERE project_id=? AND WayId=? ORDER BY Km, Picket",
            (pid, wid))
        self._load_table(self._tbl_joints_r,  JOINT_COLS,
            "SELECT * FROM tbRightIsolatedJoint WHERE project_id=? AND WayId=? ORDER BY Km, Picket",
            (pid, wid))
        self._load_table(self._tbl_pleti_l,   PLETI_COLS,
            "SELECT * FROM tbLeftPleti WHERE project_id=? AND WayId=? ORDER BY RecordNumber",
            (pid, wid))
        self._load_table(self._tbl_pleti_r,   PLETI_COLS,
            "SELECT * FROM tbRightPleti WHERE project_id=? AND WayId=? ORDER BY RecordNumber",
            (pid, wid))
        self._load_table(self._tbl_vstavki_l, VSTAVKI_COLS,
            "SELECT * FROM tbLeftVstavki WHERE project_id=? AND WayId=? ORDER BY RecordNumber",
            (pid, wid))
        self._load_table(self._tbl_vstavki_r, VSTAVKI_COLS,
            "SELECT * FROM tbRightVstavki WHERE project_id=? AND WayId=? ORDER BY RecordNumber",
            (pid, wid))
        self._load_table(self._tbl_straight,  STRAIGHT_COLS,
            "SELECT * FROM tbStraighteningData WHERE project_id=? AND WayId=? ORDER BY Km, Picket",
            (pid, wid))

    def _load_table(self, table_widget: RailwayTable,
                    col_defs: list, sql: str, params: tuple):
        loader = DataLoaderThread(self._ctx.db, sql, params)
        loader.finished.connect(
            lambda cols, rows, tw=table_widget, cd=col_defs:
                self._on_loaded(tw, cd, cols, rows)
        )
        loader.error.connect(lambda e: print(f"Load error: {e}"))
        loader.start()
        self._loaders.append(loader)

    def _on_loaded(self, table: RailwayTable, col_defs, db_cols, rows):
        table.cellChanged.disconnect()  # отключить на время загрузки
        table.fill(col_defs, db_cols, rows)
        if table is self._tbl_waydata:
            table.cellChanged.connect(self._on_cell_changed)  # переподключить
            self._count_lbl.setText(
                f"Данные профиля: {table.rowCount()} строк  |  Путь: {self._way_id}"
            )

    def _on_tab_changed(self, idx):
        self._search.clear()

    # ── Фильтрация ───────────────────────────────────────────

    def _filter_current_table(self, text: str):
        tbl = self._get_current_table()
        if not tbl:
            return
        text = text.lower()
        for row in range(tbl.rowCount()):
            match = not text
            if not match:
                for col in range(tbl.columnCount()):
                    item = tbl.item(row, col)
                    if item and text in item.text().lower():
                        match = True
                        break
            tbl.setRowHidden(row, not match)

    def _get_current_table(self) -> RailwayTable:
        idx = self._tabs.currentIndex()
        map_ = {0: self._tbl_waydata, 2: self._tbl_devices, 5: self._tbl_straight}
        return map_.get(idx)

    # ── График ───────────────────────────────────────────────

    def _show_graph(self):
        if not self._way_id:
            return
        rows = self._ctx.db.fetchall(
            "SELECT Km, Picket, Left_Rail, Right_Rail, Ground "
            "FROM tbWayData WHERE project_id=? AND WayId=? "
            "ORDER BY Km, Picket",
            (self._project_id, self._way_id)
        )
        if not rows:
            QMessageBox.information(self, "График", "Нет данных для построения графика.")
            return

        try:
            import matplotlib
            matplotlib.use("QtAgg")
            from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg
            from matplotlib.figure import Figure
            from PySide6.QtWidgets import QDialog, QVBoxLayout
        except ImportError:
            QMessageBox.critical(self, "Ошибка",
                "Для графика нужен matplotlib.\npip install matplotlib")
            return

        # Строим x как км*10 + пк для корректной оси
        xs, l_rail, r_rail, ground = [], [], [], []
        for r in rows:
            x = (r["Km"] or 0) * 10 + (r["Picket"] or 0)
            xs.append(x)
            l_rail.append(r["Left_Rail"])
            r_rail.append(r["Right_Rail"])
            ground.append(r["Ground"])

        dlg = QDialog(self)
        dlg.setWindowTitle(f"Продольный профиль — Путь {self._way_id}")
        dlg.setMinimumSize(1000, 500)
        dlg.setStyleSheet(f"background-color: {Colors.BG_BASE};")
        lay = QVBoxLayout(dlg)

        fig = Figure(figsize=(12, 5), facecolor="#0F0F0F")
        ax  = fig.add_subplot(111)
        ax.set_facecolor("#1A1A1A")

        valid_l = [(x, y) for x, y in zip(xs, l_rail) if y is not None]
        valid_r = [(x, y) for x, y in zip(xs, r_rail) if y is not None]
        valid_g = [(x, y) for x, y in zip(xs, ground) if y is not None]

        if valid_l:
            ax.plot(*zip(*valid_l), color="#FF6B00", linewidth=1.5, label="Рельс Л")
        if valid_r:
            ax.plot(*zip(*valid_r), color="#FF8533", linewidth=1.5, label="Рельс П", linestyle="--")
        if valid_g:
            ax.plot(*zip(*valid_g), color="#555", linewidth=1, label="Земля")
            if valid_l:
                ax.fill_between(
                    [x for x, _ in valid_g],
                    [y for _, y in valid_g],
                    [y for _, y in valid_l],
                    alpha=0.15, color="#FF6B00"
                )

        ax.set_xlabel("Км × 10 + Пк", color="#A0A0A0", fontsize=10)
        ax.set_ylabel("Отметка, м", color="#A0A0A0", fontsize=10)
        ax.tick_params(colors="#A0A0A0")
        ax.grid(True, alpha=0.2, color="#333")
        ax.legend(
            facecolor="#2E2E2E", edgecolor="#333",
            labelcolor="#F0F0F0", fontsize=10
        )
        for spine in ax.spines.values():
            spine.set_edgecolor("#333")

        fig.tight_layout()
        lay.addWidget(FigureCanvasQTAgg(fig))

        btn = QPushButton("Закрыть")
        btn.setFixedHeight(36)
        btn.clicked.connect(dlg.accept)
        lay.addWidget(btn)
        dlg.exec()

    # ── Экспорт Excel ─────────────────────────────────────────

    def export_excel(self):
        if not self._way_id:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from PySide6.QtWidgets import QFileDialog
        except ImportError:
            QMessageBox.critical(self, "Ошибка",
                "pip install openpyxl")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить Excel", f"Перегон_путь_{self._way_id}.xlsx",
            "Excel (*.xlsx)"
        )
        if not path:
            return

        wb  = openpyxl.Workbook()
        hdr_font  = Font(bold=True, color="F0F0F0")
        hdr_fill  = PatternFill("solid", fgColor="1E1E1E")
        hdr_align = Alignment(horizontal="center")

        # Лист: Данные профиля
        ws = wb.active
        ws.title = "Данные профиля"
        headers = [d for _, d in WAY_DATA_COLS]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(1, ci, h)
            cell.font  = hdr_font
            cell.fill  = hdr_fill
            cell.alignment = hdr_align

        for ri in range(self._tbl_waydata.rowCount()):
            for ci in range(self._tbl_waydata.columnCount()):
                item = self._tbl_waydata.item(ri, ci)
                ws.cell(ri + 2, ci + 1, item.text() if item else "")

        # Листы для кривых, устройств
        for tbl, title, cols in [
            (self._tbl_curves_l,  "Кривые Л",   CURVE_COLS),
            (self._tbl_curves_r,  "Кривые П",   CURVE_COLS),
            (self._tbl_devices,   "Устройства",  DEVICE_COLS),
            (self._tbl_straight,  "Выправка",    STRAIGHT_COLS),
        ]:
            s = wb.create_sheet(title)
            for ci, (_, h) in enumerate(cols, 1):
                cell = s.cell(1, ci, h)
                cell.font = hdr_font; cell.fill = hdr_fill
            for ri in range(tbl.rowCount()):
                for ci in range(tbl.columnCount()):
                    item = tbl.item(ri, ci)
                    s.cell(ri + 2, ci + 1, item.text() if item else "")

        wb.save(path)
        QMessageBox.information(self, "Экспорт завершён", f"Файл сохранён:\n{path}")

    def _add_row(self):
        if not self._is_editing:
            return
        tbl = self._tbl_waydata  # для активной вкладки
        row = tbl.rowCount()
        tbl.insertRow(row)
        # Заполнить project_id и WayId автоматически
        for col in range(tbl.columnCount()):
            tbl.setItem(row, col, QTableWidgetItem(""))

    def _on_cell_changed(self, row: int, col: int):
        """Подсветить изменённую ячейку."""
        item = self._tbl_waydata.item(row, col)
        if item:
            item.setBackground(
                __import__("PySide6.QtGui", fromlist=["QColor"]).QColor("#FF6B0033")
            )
    
    def _save_changes(self):
        if not self._is_editing or not self._way_id:
            return
    
        from PySide6.QtWidgets import QMessageBox
        
        saved  = 0
        errors = 0
        col_defs = WAY_DATA_COLS  # список (field, label)
    
        for row in range(self._tbl_waydata.rowCount()):
            # Получить RecordNumber из первой колонки
            record_item = self._tbl_waydata.item(row, 0)
            if not record_item or not record_item.text():
                continue
            
            try:
                record_num = int(record_item.text())
            except ValueError:
                continue
    
            # Собрать все значения строки
            updates = {}
            for col_idx, (field, _) in enumerate(col_defs):
                if field == "RecordNumber":
                    continue
                item = self._tbl_waydata.item(row, col_idx)
                if item:
                    val = item.text().strip() or None
                    updates[field] = val
    
            if not updates:
                continue
    
            # Построить UPDATE запрос
            sets   = ", ".join(f"{k}=?" for k in updates.keys())
            values = list(updates.values()) + [record_num, self._project_id]
            
            try:
                self._ctx.db.execute(
                    f"UPDATE tbWayData SET {sets} "
                    f"WHERE RecordNumber=? AND project_id=?",
                    values
                )
                saved += 1
            except Exception as e:
                errors += 1
                print(f"Save error row {row}: {e}")
    
        self._ctx.db.commit()
    
        # Убрать подсветку
        for row in range(self._tbl_waydata.rowCount()):
            for col in range(self._tbl_waydata.columnCount()):
                item = self._tbl_waydata.item(row, col)
                if item:
                    item.setBackground(
                        __import__("PySide6.QtGui", fromlist=["QColor"]).QColor(0, 0, 0, 0)
                    )
    
        QMessageBox.information(
            self, "Сохранено",
            f"Сохранено строк: {saved}\nОшибок: {errors}"
        )

    def _add_way(self):
        from ui.dialogs.add_way_dialog import AddWayDialog
        dlg = AddWayDialog(self._ctx, self._project_id, self)
        if dlg.exec():
            self._load_ways()