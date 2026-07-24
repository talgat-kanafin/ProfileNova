"""
ui/screens/access_way_page.py
==============================
Раздел "Подъездные пути" — tbAccessWay* таблицы.

Вкладки:
  1. Основные данные  (tbAccessWayData)
  2. Кривые           (tbAccessWayCurve)
  3. Устройства       (tbAccessWayDevice)
"""

import os, sys
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QTabWidget, QPushButton, QComboBox, QFrame,
    QMessageBox, QLineEdit
)

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from ui.theme import Colors, Spacing, Sizes, Typography
from ui.screens.mainline_page import RailwayTable, DataLoaderThread


ACCESS_DATA_COLS = [
    ("RecordNumber", "№"),
    ("Picket",       "Пк"),
    ("Meter",        "Метр"),
    ("Ground",       "Земля"),
    ("Rail",         "Рельс"),
    ("SUklon",       "Уклон"),
    ("GroundEdge",   "Бровка"),
    ("BallastInfo",  "Балласт"),
    ("TieInfo",      "Шпалы"),
]

ACCESS_CURVE_COLS = [
    ("RecordNumber", "№"),
    ("AngleDegree",  "Угол°"),
    ("AngleMinute",  "Угол'"),
    ("AngleSecond",  "Угол''"),
    ("Radius",       "Радиус"),
    ("Tangent",      "Тангенс"),
    ("Tangent2",     "Тангенс2"),
    ("Curvature",    "Кривизна"),
    ("Length",       "Длина"),
    ("Length2",      "Длина2"),
    ("startPk",      "От Пк"),
    ("startM",       "От М"),
    ("endPk",        "До Пк"),
    ("endM",         "До М"),
    ("B",            "B"),
]

ACCESS_DEVICE_COLS = [
    ("RecordNumber", "№"),
    ("Picket",       "Пк"),
    ("Meter",        "Метр"),
    ("DeviceId",     "Устройство"),
    ("Note",         "Примечание"),
]


class AccessWayPage(QWidget):

    def __init__(self, ctx, project_id: int, is_editing: bool, parent=None):
        super().__init__(parent)
        self._ctx        = ctx
        self._project_id = project_id
        self._is_editing = is_editing
        self._way_id     = None
        self._loaders    = []

        self._build_ui()
        self._load_ways()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # Тулбар
        tb = QFrame()
        tb.setFixedHeight(Sizes.TOOLBAR_HEIGHT)
        tb.setStyleSheet(f"""
            background-color: {Colors.BG_SURFACE};
            border-bottom: 1px solid {Colors.BORDER};
        """)
        tb_layout = QHBoxLayout(tb)
        tb_layout.setContentsMargins(Spacing.LG, 0, Spacing.LG, 0)
        tb_layout.setSpacing(Spacing.MD)

        tb_layout.addWidget(QLabel("Подъездной путь:"))
        self._way_combo = QComboBox()
        self._way_combo.setFixedWidth(300)
        self._way_combo.setFixedHeight(32)
        self._way_combo.currentIndexChanged.connect(self._on_way_changed)
        tb_layout.addWidget(self._way_combo)
        btn_add_way = QPushButton("+")
        btn_add_way.setFixedSize(32, 32)
        btn_add_way.setFont(Typography.get_font(16, bold=True))
        btn_add_way.setToolTip("Добавить подъездной путь")
        btn_add_way.setStyleSheet(f"""
            QPushButton {{
                background-color: {Colors.BG_ELEVATED};
                color: {Colors.TEXT_PRIMARY};
                border: 1px solid {Colors.BORDER};
                border-radius: 4px;
            }}
            QPushButton:hover {{
                background-color: {Colors.ACCENT};
                color: white;
                border-color: {Colors.ACCENT};
            }}
        """)
        btn_add_way.clicked.connect(self._add_access_way)
        tb_layout.addWidget(btn_add_way)

        tb_layout.addStretch()

        self._search = QLineEdit()
        self._search.setPlaceholderText("Поиск...")
        self._search.setFixedWidth(180)
        self._search.setFixedHeight(32)
        self._search.textChanged.connect(self._filter_current)
        tb_layout.addWidget(self._search)

        btn_export = QPushButton("📊 Excel")
        btn_export.setFixedHeight(32)
        btn_export.clicked.connect(self.export_excel)
        tb_layout.addWidget(btn_export)

        root.addWidget(tb)

        # Счётчик
        self._count_lbl = QLabel("")
        self._count_lbl.setStyleSheet(f"""
            background-color: {Colors.BG_SURFACE};
            color: {Colors.TEXT_SECONDARY};
            padding: 3px {Spacing.LG}px;
            font-size: 11px;
            border-bottom: 1px solid {Colors.BORDER};
        """)
        root.addWidget(self._count_lbl)

        # Вкладки
        self._tabs = QTabWidget()
        self._tabs.setStyleSheet(f"""
            QTabWidget::pane {{ border: none; background-color: {Colors.BG_BASE}; }}
            QTabBar::tab {{
                background-color: {Colors.BG_SURFACE};
                color: {Colors.TEXT_SECONDARY};
                padding: 8px 18px; border: none;
                border-bottom: 2px solid transparent; font-size: 12px;
            }}
            QTabBar::tab:selected {{
                color: {Colors.TEXT_PRIMARY};
                border-bottom: 2px solid {Colors.ACCENT};
            }}
            QTabBar::tab:hover {{
                color: {Colors.TEXT_PRIMARY};
                background-color: {Colors.BG_ELEVATED};
            }}
        """)

        self._tbl_data    = RailwayTable(self._is_editing)
        self._tbl_curves  = RailwayTable(self._is_editing)
        self._tbl_devices = RailwayTable(self._is_editing)
        
        self._tbl_data.cellChanged.connect(self._on_cell_changed)

        self._tabs.addTab(self._tbl_data,    "Основные данные")
        self._tabs.addTab(self._tbl_curves,  "Кривые")
        self._tabs.addTab(self._tbl_devices, "Устройства")
        self._tabs.currentChanged.connect(lambda _: self._search.clear())

        root.addWidget(self._tabs, 1)

    # ── Загрузка ─────────────────────────────────────────────

    def _load_ways(self):
        rows = self._ctx.db.fetchall(
            "SELECT WayId, Name, stPk, enPk FROM tbAccessWayInfo "
            "WHERE project_id=? ORDER BY WayId",
            (self._project_id,)
        )
        self._way_combo.blockSignals(True)
        self._way_combo.clear()

        if not rows:
            self._count_lbl.setText("Нет подъездных путей в проекте")
            self._way_combo.blockSignals(False)
            return

        for r in rows:
            name = r.get("Name") or f"Путь {r['WayId']}"
            pk   = ""
            if r.get("stPk") is not None and r.get("enPk") is not None:
                pk = f"  (пк {r['stPk']} — {r['enPk']})"
            self._way_combo.addItem(name + pk, r["WayId"])

        self._way_combo.blockSignals(False)
        self._way_id = self._way_combo.currentData()
        self._load_data()

    def _on_way_changed(self):
        self._way_id = self._way_combo.currentData()
        if self._way_id is not None:
            self._load_data()

    def _load_data(self):
        if not self._way_id:
            return
        pid = self._project_id
        wid = self._way_id

        self._load_table(
            self._tbl_data, ACCESS_DATA_COLS,
            "SELECT * FROM tbAccessWayData "
            "WHERE project_id=? AND AccessWayId=? ORDER BY Picket, Meter",
            (pid, wid)
        )
        self._load_table(
            self._tbl_curves, ACCESS_CURVE_COLS,
            "SELECT * FROM tbAccessWayCurve "
            "WHERE project_id=? AND AccessWayId=? ORDER BY RecordNumber",
            (pid, wid)
        )
        self._load_table(
            self._tbl_devices, ACCESS_DEVICE_COLS,
            "SELECT d.RecordNumber, d.Picket, d.Meter, "
            "       dev.DeviceName as DeviceId, d.Note "
            "FROM tbAccessWayDevice d "
            "LEFT JOIN tbDevice dev ON dev.DeviceId = d.DeviceId "
            "WHERE d.project_id=? AND d.AccessWayId=? "
            "ORDER BY d.Picket, d.Meter",
            (pid, wid)
        )

    def _load_table(self, table_widget, col_defs, sql, params):
        loader = DataLoaderThread(self._ctx.db, sql, params)
        loader.finished.connect(
            lambda cols, rows, tw=table_widget, cd=col_defs:
                self._on_loaded(tw, cd, cols, rows)
        )
        loader.error.connect(lambda e: print(f"AccessWay load error: {e}"))
        loader.start()
        self._loaders.append(loader)

    def _on_loaded(self, table, col_defs, db_cols, rows):
        try:
            table.cellChanged.disconnect()
        except Exception:
            pass
        table.fill(col_defs, db_cols, rows)
        if table is self._tbl_data:
            table.cellChanged.connect(self._on_cell_changed)
            way_name = self._way_combo.currentText()
            self._count_lbl.setText(f"{way_name}  |  Строк: {table.rowCount()}")
  
    # ── Фильтрация ───────────────────────────────────────────

    def _filter_current(self, text: str):
        idx = self._tabs.currentIndex()
        tbl = [self._tbl_data, self._tbl_curves, self._tbl_devices][idx]
        text = text.lower()
        for row in range(tbl.rowCount()):
            match = not text
            if not match:
                for col in range(tbl.columnCount()):
                    item = tbl.item(row, col)
                    if item and text in item.text().lower():
                        match = True
                        break
            tbl.setRowHidden(row, not match)

    # ── Экспорт ──────────────────────────────────────────────

    def export_excel(self):
        if not self._way_id:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from PySide6.QtWidgets import QFileDialog
        except ImportError:
            QMessageBox.critical(self, "Ошибка", "pip install openpyxl")
            return

        way_name = self._way_combo.currentText().split("  ")[0]
        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить Excel",
            f"Подъездной_{way_name}.xlsx",
            "Excel (*.xlsx)"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        hdr_font = Font(bold=True, color="F0F0F0")
        hdr_fill = PatternFill("solid", fgColor="1E1E1E")

        for tbl, title, cols in [
            (self._tbl_data,    "Основные данные", ACCESS_DATA_COLS),
            (self._tbl_curves,  "Кривые",          ACCESS_CURVE_COLS),
            (self._tbl_devices, "Устройства",       ACCESS_DEVICE_COLS),
        ]:
            ws = wb.active if title == "Основные данные" else wb.create_sheet(title)
            ws.title = title
            for ci, (_, h) in enumerate(cols, 1):
                cell = ws.cell(1, ci, h)
                cell.font = hdr_font
                cell.fill = hdr_fill
            for ri in range(tbl.rowCount()):
                for ci in range(tbl.columnCount()):
                    item = tbl.item(ri, ci)
                    ws.cell(ri + 2, ci + 1, item.text() if item else "")

        wb.save(path)
        QMessageBox.information(self, "Экспорт завершён", f"Файл сохранён:\n{path}")

    def _on_cell_changed(self, row: int, col: int):
        from PySide6.QtGui import QColor
        item = self._tbl_data.item(row, col)
        if item:
            item.setBackground(QColor("#FF6B0033"))

    def _add_access_way(self):
        from PySide6.QtWidgets import QDialog, QVBoxLayout, QFormLayout
        from PySide6.QtWidgets import QSpinBox, QLineEdit, QDialogButtonBox, QLabel
        
        dlg = QDialog(self)
        dlg.setWindowTitle("Добавить подъездной путь")
        dlg.setFixedSize(360, 260)
        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(16, 16, 16, 16)
        
        form = QFormLayout()
            
        way_id_spin = QSpinBox()
        way_id_spin.setRange(1, 9999)
        row = self._ctx.db.fetchone(
            "SELECT MAX(WayId) as m FROM tbAccessWayInfo WHERE project_id=?",
            (self._project_id,)
        )
        way_id_spin.setValue((row["m"] or 0) + 1 if row else 1)
        way_id_spin.setFixedHeight(36)
        form.addRow("ID пути *:", way_id_spin)
        
        name_edit = QLineEdit()
        name_edit.setFixedHeight(36)
        name_edit.setPlaceholderText("Название подъездного пути")
        form.addRow("Название:", name_edit)
        
        st_pk = QSpinBox()
        st_pk.setRange(0, 9999)
        st_pk.setFixedHeight(36)
        form.addRow("Начало Пк:", st_pk)
        
        en_pk = QSpinBox()
        en_pk.setRange(0, 9999)
        en_pk.setFixedHeight(36)
        form.addRow("Конец Пк:", en_pk)
        
        layout.addLayout(form)
        layout.addStretch()
        
        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save |
            QDialogButtonBox.StandardButton.Cancel
        )
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        layout.addWidget(btns)
        
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        
        self._ctx.db.execute(
            "INSERT INTO tbAccessWayInfo (WayId, project_id, stPk, enPk, Name) VALUES (?,?,?,?,?)",
            (way_id_spin.value(), self._project_id,
             st_pk.value(), en_pk.value(),
             name_edit.text().strip() or None)
        )
        self._ctx.db.commit()
        self._load_ways()